from flask import Flask, request, jsonify, send_file, send_from_directory, render_template, session
from flask_cors import CORS
from datetime import datetime, timedelta
from functools import wraps
import database as db
import os
import smtplib
from email.message import EmailMessage
from werkzeug.utils import secure_filename
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("python-dotenv not found. Assuming environment variables are provided by the host.")

app = Flask(__name__, static_folder='www', template_folder='www')
app.secret_key = os.environ.get('SECRET_KEY', 'church-attendance-secret-key-2024')
app.permanent_session_lifetime = timedelta(days=30)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024
CORS(app, supports_credentials=True)

ROLE_LEVELS = {
    'user': 1,
    'rewards': 1,
    'sub_admin': 2,
    'admin': 3,
}
ALLOWED_ANNOUNCEMENT_EXTENSIONS = {
    '.pdf', '.png', '.jpg', '.jpeg', '.gif', '.webp',
    '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx',
}


def current_user():
    """Return the logged-in user, refreshed from the database."""
    username = session.get('username')
    if not username:
        return None

    user = db.get_user_by_username(username)
    if not user:
        session.clear()
        return None

    if user.get('username', '').upper() == 'STJOHN':
        user['role'] = 'admin'
        user['class_name'] = 'all'
        user['assigned_grades'] = ''
        user['assigned_sections'] = ''

    session['username'] = user['username']
    session['role'] = user.get('role', 'user')
    session['class_name'] = user.get('class_name', 'all')
    session['assigned_grades'] = user.get('assigned_grades', '')
    session['assigned_sections'] = user.get('assigned_sections', '')
    return {
        'username': user['username'],
        'role': user.get('role', 'user'),
        'class_name': user.get('class_name', 'all'),
        'assigned_grades': user.get('assigned_grades', ''),
        'assigned_sections': user.get('assigned_sections', ''),
    }


def _norm_grade(grade):
    value = db.normalize_grade_value(grade)
    return str(value).strip().lower() if value is not None else ''


def _assigned_grade_set(user):
    text = db.normalize_assigned_grades(user.get('assigned_grades', ''))
    return {_norm_grade(g) for g in text.split(',') if _norm_grade(g)}


def _norm_gender(gender):
    return db.normalize_gender_value(gender).strip().lower()


def _assigned_section_set(user):
    text = db.normalize_assigned_sections(user.get('assigned_sections', ''))
    return {_norm_gender(s) for s in text.split(',') if _norm_gender(s)}


def _student_class(student):
    return (student.get('class_name') or db.grade_to_class_name(student.get('grade')) or '').strip()


def _user_section_allows(user, student):
    allowed_sections = _assigned_section_set(user)
    if not allowed_sections:
        return True
    return _norm_gender(student.get('gender')) in allowed_sections


def can_user_access_student(user, student):
    """Apply admin/sub-admin/servant visibility rules to one student dict."""
    if not user:
        return False
    role = user.get('role')
    if role == 'admin':
        return True

    student_grade = _norm_grade(student.get('grade'))
    student_class = _student_class(student).lower()
    user_class = (user.get('class_name') or 'all').strip()

    if role == 'sub_admin':
        if user_class == 'all':
            return True
        allowed = {_norm_grade(g) for g in db.class_grade_values(user_class)}
        return student_class == user_class.lower() or student_grade in allowed

    assigned_grades = _assigned_grade_set(user)
    if role == 'rewards':
        if assigned_grades:
            return student_grade in assigned_grades and _user_section_allows(user, student)
        return _user_section_allows(user, student)

    if assigned_grades:
        return student_grade in assigned_grades and _user_section_allows(user, student)

    if user_class and user_class != 'all':
        allowed = {_norm_grade(g) for g in db.class_grade_values(user_class)}
        return (student_class == user_class.lower() or student_grade in allowed) and _user_section_allows(user, student)

    return False


def filter_students_for_user(students, user=None):
    user = user or current_user()
    return [student for student in students if can_user_access_student(user, student)]


def get_student_if_allowed(student_id, user=None):
    user = user or current_user()
    try:
        student = db.get_student_details(student_id)
    except Exception:
        student = None
    if not student or not can_user_access_student(user, student):
        return None
    return student


def resolve_student_id_from_record(user, record):
    """Resolve local app student records to the correct server student id."""
    visible_student = {
        'id': record.get('student_id'),
        'name': record.get('student_name') or record.get('name'),
        'grade': record.get('grade'),
        'gender': record.get('gender'),
        'servant': record.get('servant') or user.get('username', ''),
        'class_name': db.grade_to_class_name(record.get('grade')),
    }

    if visible_student.get('grade') not in (None, '') or visible_student.get('gender'):
        if not can_user_access_student(user, visible_student):
            return None

    student_id = record.get('student_id')
    existing = None
    if student_id:
        try:
            existing = db.get_student_details(student_id)
        except Exception:
            existing = None

    if existing and can_user_access_student(user, existing):
        payload_name = (visible_student.get('name') or '').strip().lower()
        same_name = not payload_name or (existing.get('name') or '').strip().lower() == payload_name
        same_grade = not visible_student.get('grade') or _norm_grade(existing.get('grade')) == _norm_grade(visible_student.get('grade'))
        same_gender = not visible_student.get('gender') or _norm_gender(existing.get('gender')) == _norm_gender(visible_student.get('gender'))
        if same_name and same_grade and same_gender:
            return student_id

    if visible_student.get('name'):
        return db.find_or_create_student_from_record(record, user.get('username', ''))

    return None


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            return jsonify({'error': 'Authentication required'}), 401
        return fn(*args, **kwargs)
    return wrapper


def role_required(*roles):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                return jsonify({'error': 'Authentication required'}), 401
            if user.get('role') not in roles:
                return jsonify({'error': 'Permission denied'}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def min_role_required(role):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                return jsonify({'error': 'Authentication required'}), 401
            current_level = ROLE_LEVELS.get(user.get('role'), 0)
            required_level = ROLE_LEVELS.get(role, 0)
            if current_level < required_level:
                return jsonify({'error': 'Permission denied'}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator

# Email Configuration (Loaded from .env file)
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "") 
SENDER_PASSWORD = os.environ.get("SENDER_PASSWORD", "") # e.g. Gmail App Password

# Always run init_db to ensure all tables exist
db.init_db()

# Import from Excel if DB was newly created and Excel file exists
excel_path = os.environ.get("INITIAL_EXCEL_IMPORT_PATH", "")
if not os.path.exists(db.DB_FILE) or os.path.getsize(db.DB_FILE) < 1000:
    if os.path.exists(excel_path):
        db.import_from_excel(excel_path)
        print("[SUCCESS] Database initialized and data imported from Excel")
    else:
        print("[WARNING] Excel file not found, starting with empty database")
else:
    print("[SUCCESS] Database loaded successfully")

@app.route('/')
def index():
    """Serve the main application page from the mobile www folder."""
    response = send_file(os.path.join('www', 'index.html'), max_age=0)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response

@app.route('/<path:filename>')
def serve_static(filename):
    """Serve static files from the mobile www folder."""
    response = send_from_directory('www', filename, max_age=0)
    if filename.endswith(('.html', '.js', '.css')):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response

@app.route('/api/send-email', methods=['POST'])
@min_role_required('sub_admin')
def send_email():
    """Send an automated background email to selected recipients."""
    data = request.json
    subject = data.get('subject')
    message_body = data.get('message')
    recipients = data.get('recipients', []) # List of email addresses

    if not subject or not message_body or not recipients:
        return jsonify({'error': 'Missing subject, message, or recipients'}), 400
    
    if not SENDER_EMAIL or not SENDER_PASSWORD:
        return jsonify({'error': 'Email server not configured. Please set SENDER_EMAIL and SENDER_PASSWORD.'}), 500

    try:
        # Create the email
        msg = EmailMessage()
        msg.set_content(message_body)
        msg['Subject'] = subject
        msg['From'] = SENDER_EMAIL
        msg['Bcc'] = ", ".join(recipients) # Send as BCC for privacy

        # Connect to SMTP server and send
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)

        return jsonify({'success': True, 'count': len(recipients)})
    except smtplib.SMTPAuthenticationError:
        return jsonify({'error': 'SMTP Authentication failed. Check your App Password.'}), 500
    except Exception as e:
        return jsonify({'error': f'Failed to send email: {str(e)}'}), 500

@app.route('/api/servants', methods=['GET'])
@login_required
def get_servants():
    """Get list of all servants."""
    try:
        servants = db.get_servants()
        return jsonify({'servants': servants})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/filters', methods=['GET'])
@login_required
def get_filters():
    """Get available filters (grades, genders) for a specific servant."""
    servant = request.args.get('servant')
    if not servant:
        return jsonify({'error': 'Servant parameter required'}), 400
    
    try:
        filters = db.get_filters(servant)
        return jsonify(filters)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/students', methods=['GET'])
@login_required
def get_students():
    """Get students with optional filters."""
    servant = request.args.get('servant')
    grade = request.args.get('grade', type=int)
    gender = request.args.get('gender')
    
    try:
        user = current_user()
        students = db.get_students(servant, grade, gender)
        students = filter_students_for_user(students, user)
        return jsonify({'students': students})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/attendance', methods=['POST'])
@login_required
def save_attendance():
    """Save attendance record."""
    data = request.json
    
    student_id = data.get('student_id')
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    status = data.get('status')  # 'present' or 'absent'
    
    if not all([student_id, status]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    try:
        if not get_student_if_allowed(student_id):
            return jsonify({'error': 'Permission denied for this student'}), 403
        db.save_attendance(student_id, date, status)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student', methods=['POST'])
@min_role_required('sub_admin')
def add_student():
    """Add a new student."""
    data = request.json
    
    name = data.get('name')
    grade = data.get('grade')
    gender = data.get('gender')
    servant = data.get('servant')
    
    if not all([name, grade, gender, servant]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    try:
        student_id = db.add_student(
            name, grade, gender, servant,
            phone=data.get('phone', ''),
            parent_phone=data.get('parent_phone', ''),
            dob=data.get('dob', ''),
            address=data.get('address', ''),
            comments=data.get('comments', ''),
            pictures=data.get('pictures', ''),
            last_call=data.get('last_call', '')
        )
        return jsonify({'success': True, 'student_id': student_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/<int:student_id>', methods=['PUT'])
@min_role_required('sub_admin')
def update_student(student_id):
    """Update student information."""
    data = request.json
    
    try:
        if not get_student_if_allowed(student_id):
            return jsonify({'error': 'Permission denied for this student'}), 403
        db.update_student(student_id, **data)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/<int:student_id>', methods=['DELETE'])
@min_role_required('sub_admin')
def delete_student(student_id):
    """Delete a student."""
    try:
        if not get_student_if_allowed(student_id):
            return jsonify({'error': 'Permission denied for this student'}), 403
        db.delete_student(student_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/history/<int:student_id>', methods=['GET'])
@login_required
def get_history(student_id):
    """Get attendance history for a student."""
    try:
        if not get_student_if_allowed(student_id):
            return jsonify({'error': 'Permission denied for this student'}), 403
        history = db.get_attendance_history(student_id)
        return jsonify({'history': history})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics', methods=['GET'])
@login_required
def get_analytics():
    """Get analytics data for dashboard."""
    try:
        analytics = db.get_analytics()
        return jsonify(analytics)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export', methods=['GET'])
@min_role_required('sub_admin')
def export_excel():
    """Export data to Excel file."""
    try:
        output_path = 'export_attendance_data.xlsx'
        db.export_to_excel(output_path)
        return send_file(output_path, as_attachment=True, download_name=f'attendance_export_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/servant', methods=['POST'])
@role_required('admin')
def add_servant():
    """Add a new servant."""
    data = request.json
    name = data.get('name')
    phone = data.get('phone', '')
    
    if not name:
        return jsonify({'error': 'Servant name is required'}), 400
    
    try:
        # Add servant to the servants table (or create one if needed)
        db.add_servant(name, phone)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/students-by-servant', methods=['GET'])
@login_required
def get_students_by_servant():
    """Get all students assigned to a specific servant."""
    servant = request.args.get('servant')
    if not servant:
        return jsonify({'error': 'Servant parameter required'}), 400
    
    try:
        user = current_user()
        students = db.get_students_by_servant(servant)
        students = filter_students_for_user(students, user)
        return jsonify({'students': students})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/student/<int:student_id>/details', methods=['GET'])
@login_required
def get_student_details(student_id):
    """Get full student details including attendance history and notes."""
    try:
        details = db.get_student_details(student_id)
        if details and can_user_access_student(current_user(), details):
            return jsonify(details)
        if details:
            return jsonify({'error': 'Permission denied for this student'}), 403
        return jsonify({'error': 'Student not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/notes', methods=['POST'])
@login_required
def add_note():
    """Add a note to a student."""
    data = request.json
    student_id = data.get('student_id')
    note_text = data.get('note_text')
    created_by = data.get('created_by', '')
    
    if not student_id or not note_text:
        return jsonify({'error': 'student_id and note_text are required'}), 400
    
    try:
        if not get_student_if_allowed(student_id):
            return jsonify({'error': 'Permission denied for this student'}), 403
        note_id = db.add_note(student_id, note_text, created_by)
        return jsonify({'success': True, 'note_id': note_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/notes/<int:student_id>', methods=['GET'])
@login_required
def get_notes(student_id):
    """Get all notes for a student."""
    try:
        if not get_student_if_allowed(student_id):
            return jsonify({'error': 'Permission denied for this student'}), 403
        notes = db.get_notes(student_id)
        return jsonify({'notes': notes})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/filters-all', methods=['GET'])
@login_required
def get_filters_all():
    """Get all available grades and genders (no servant filter)."""
    try:
        students = filter_students_for_user(db.get_students())
        grades = sorted(
            {s.get('grade') for s in students if s.get('grade') not in (None, '')},
            key=lambda g: str(g)
        )
        genders = sorted({s.get('gender') for s in students if s.get('gender')})
        return jsonify({'grades': grades, 'genders': genders})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-excel', methods=['POST'])
@min_role_required('sub_admin')
def upload_excel():
    """Upload and import a new Excel file."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file type. Please upload an Excel file.'}), 400
    
    try:
        # Save the file temporarily
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            result = db.import_from_excel_upload(tmp.name)
            os.unlink(tmp.name)  # Delete temp file
        
        if result['success']:
            return jsonify({
                'success': True, 
                'message': f"Imported {result['added']} new kids, updated {result['updated']} existing kids"
            })
        else:
            return jsonify({'error': result.get('error', 'Import failed')}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-attendance', methods=['POST'])
@login_required
def batch_attendance():
    """Save multiple attendance records at once and award points."""
    data = request.json
    records = data.get('records', [])
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))

    try:
        user = current_user()
        saved_count = 0
        for record in records:
            status = record.get('status')
            student_id = resolve_student_id_from_record(user, record)
            if student_id and status:
                liturgy   = int(record.get('liturgy', 0))
                tonia     = int(record.get('tonia', 0))
                confession= int(record.get('confession', 0))
                bible_p   = int(record.get('bible_prayer', 0))
                questions = int(record.get('questions', 0))
                is_present = 1 if status == 'present' else 0

                # Save attendance with extra fields
                conn = __import__('sqlite3').connect(db.DB_FILE)
                conn.execute('''
                    INSERT OR REPLACE INTO attendance
                    (student_id, date, status, liturgy, tonia, confession, bible_prayer, questions)
                    VALUES (?,?,?,?,?,?,?,?)
                ''', (student_id, date, status, liturgy, tonia, confession, bible_p, questions))
                conn.commit()
                conn.close()

                # Award points
                db.calculate_and_award_points(student_id, date, is_present,
                                              liturgy, tonia, confession, bible_p, questions)
                saved_count += 1
            elif status:
                return jsonify({'error': 'Permission denied for one or more students'}), 403

        return jsonify({'success': True, 'count': saved_count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== POINTS ENDPOINTS ====================

@app.route('/api/points/all', methods=['GET'])
@login_required
def get_all_points():
    """Get points leaderboard for all students."""
    try:
        user = current_user()
        servant = request.args.get('servant')
        students = db.get_all_students_points()
        if servant and servant != 'all':
            class_grades = {_norm_grade(g) for g in db.class_grade_values(servant)}
            students = [
                s for s in students
                if s.get('servant') == servant
                or s.get('class_name') == servant
                or _norm_grade(s.get('grade')) in class_grades
            ]
        students = filter_students_for_user(students, user)
        return jsonify({'students': students})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/points/<int:student_id>', methods=['GET'])
@login_required
def get_student_points_route(student_id):
    try:
        if not get_student_if_allowed(student_id):
            return jsonify({'error': 'Permission denied for this student'}), 403
        data = db.get_student_points(student_id)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/points/add', methods=['POST'])
@role_required('admin', 'sub_admin', 'rewards')
def add_points():
    """Manually add or deduct points."""
    data = request.json
    student_id = data.get('student_id')
    points = int(data.get('points', 0))
    reason = data.get('reason', 'Manual adjustment')
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    if not student_id:
        return jsonify({'error': 'student_id required'}), 400
    try:
        if not get_student_if_allowed(student_id):
            return jsonify({'error': 'Permission denied for this student'}), 403
        db.add_manual_points(student_id, points, reason, date)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/points/clear', methods=['POST'])
@min_role_required('sub_admin')
def clear_points():
    """Clear (redeem) points for all or a class. History is preserved."""
    data = request.json or {}
    servant = data.get('servant')
    try:
        count = db.clear_points_for_redemption(servant if servant and servant != 'all' else None)
        return jsonify({'success': True, 'cleared_count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== EFTIKAD ENDPOINTS ====================

@app.route('/api/eftikad', methods=['POST'])
@login_required
def submit_eftikad():
    data = request.json
    servant_username = data.get('servant_username', '')
    user = current_user()
    student_id = resolve_student_id_from_record(user, data)
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    if not student_id:
        return jsonify({'error': 'student_id required'}), 400
    try:
        eid = db.submit_eftikad(
            servant_username, student_id, date,
            int(data.get('whatsapp_sent', 0)),
            int(data.get('called', 0)),
            int(data.get('visited', 0)),
            data.get('notes', '')
        )
        return jsonify({'success': True, 'eftikad_id': eid})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/eftikad', methods=['GET'])
@login_required
def get_eftikad():
    servant = request.args.get('servant')
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    try:
        user = current_user()
        rows = db.get_eftikad(servant, date_from, date_to)
        rows = [row for row in rows if can_user_access_student(user, row)]
        return jsonify({'eftikad': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== ANNOUNCEMENTS ENDPOINTS ====================

@app.route('/api/announcements', methods=['GET'])
@login_required
def get_announcements():
    try:
        rows = db.get_announcements()
        return jsonify({'announcements': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/announcements', methods=['POST'])
@min_role_required('sub_admin')
def post_announcement():
    title = request.form.get('title', '')
    body = request.form.get('body', '')
    created_by = request.form.get('created_by', '')
    attachment_path = ''

    if 'attachment' in request.files:
        file = request.files['attachment']
        if file and file.filename:
            _, ext = os.path.splitext(file.filename)
            if ext.lower() not in ALLOWED_ANNOUNCEMENT_EXTENSIONS:
                return jsonify({'error': 'Unsupported attachment type'}), 400
            uploads_dir = os.path.join(os.path.dirname(__file__), 'uploads')
            os.makedirs(uploads_dir, exist_ok=True)
            safe_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
            save_path = os.path.join(uploads_dir, safe_name)
            file.save(save_path)
            attachment_path = f'/uploads/{safe_name}'

    if not title or not body:
        return jsonify({'error': 'title and body required'}), 400

    try:
        aid = db.add_announcement(title, body, created_by, attachment_path)
        return jsonify({'success': True, 'id': aid})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/announcements/<int:ann_id>', methods=['DELETE'])
@min_role_required('sub_admin')
def delete_announcement(ann_id):
    try:
        ok = db.delete_announcement(ann_id)
        return jsonify({'success': ok})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    uploads_dir = os.path.join(os.path.dirname(__file__), 'uploads')
    return send_from_directory(uploads_dir, filename)


# ==================== BIRTHDAYS ENDPOINT ====================

@app.route('/api/birthdays', methods=['GET'])
@login_required
def get_birthdays():
    days = int(request.args.get('days', 30))
    try:
        birthdays = db.get_upcoming_birthdays(days)
        birthdays = filter_students_for_user(birthdays)
        return jsonify({'birthdays': birthdays})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== BIBLE TRACKING ENDPOINTS ====================

@app.route('/api/bible/<int:student_id>', methods=['GET'])
@login_required
def get_bible(student_id):
    try:
        data = db.get_bible_tracking(student_id)
        return jsonify({'tracking': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/bible', methods=['POST'])
@login_required
def update_bible():
    data = request.json
    student_id = data.get('student_id')
    week_start = data.get('week_start_date')
    days_read = int(data.get('days_read', 0))
    if not student_id or not week_start:
        return jsonify({'error': 'student_id and week_start_date required'}), 400
    try:
        db.update_bible_tracking(student_id, week_start, days_read)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/bible/all', methods=['GET'])
@login_required
def get_bible_all():
    servant = request.args.get('servant')
    try:
        rows = db.get_bible_tracking_all(servant)
        return jsonify({'students': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== CHARTS ENDPOINT ====================

@app.route('/api/charts/attendance', methods=['GET'])
@login_required
def get_chart_data():
    servant = request.args.get('servant')
    year = request.args.get('year')
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    try:
        user = current_user()
        grades = None
        sections = None
        servant_filter = servant
        if user and user.get('role') != 'admin':
            sections = user.get('assigned_sections', '')
            if user.get('role') in ('user', 'rewards'):
                grades = user.get('assigned_grades', '')
                servant_filter = None
                if not grades and user.get('class_name') and user.get('class_name') != 'all':
                    grades = db.class_grade_values(user.get('class_name'))
            elif user.get('class_name') and user.get('class_name') != 'all':
                servant_filter = user.get('class_name')
        data = db.get_class_attendance_stats(servant_filter, year, grades, sections, date_from, date_to)
        return jsonify({'stats': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/login', methods=['POST'])
def login():
    """Verify user credentials."""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    
    try:
        user = db.check_user(username, password)
        if user:
            session.clear()
            session.permanent = True
            session['username'] = user['username']
            session['role'] = user['role']
            session['class_name'] = user.get('class_name', 'all')
            session['assigned_grades'] = user.get('assigned_grades', '')
            session['assigned_sections'] = user.get('assigned_sections', '')
            return jsonify({
                'success': True,
                'role': user['role'],
                'username': user['username'],
                'class_name': user.get('class_name', 'all'),
                'assigned_grades': user.get('assigned_grades', ''),
                'assigned_sections': user.get('assigned_sections', '')
            })
        else:
            return jsonify({'success': False, 'error': 'Invalid username or password'}), 401
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/session', methods=['GET'])
def get_session():
    """Return the currently authenticated user, if any."""
    user = current_user()
    if not user:
        return jsonify({'authenticated': False}), 401
    return jsonify({'authenticated': True, **user})


@app.route('/api/logout', methods=['POST'])
def logout():
    """Clear the current login session."""
    session.clear()
    return jsonify({'success': True})


# ==================== ADMIN USER MANAGEMENT ====================

@app.route('/admin')
def admin_dashboard():
    """Serve the admin dashboard page."""
    response = send_file(os.path.join('www', 'admin.html'), max_age=0)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


@app.route('/api/admin/users', methods=['GET'])
@role_required('admin')
def get_users():
    """Get all users (admin only - frontend enforces auth)."""
    try:
        users = db.get_all_users()
        return jsonify({'users': users})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users', methods=['POST'])
@role_required('admin')
def create_user():
    """Create a new user."""
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'user')
    class_name = (data.get('class_name') or '').strip()
    assigned_grades = data.get('assigned_grades', [])
    assigned_sections = data.get('assigned_sections', [])
    if not class_name:
        class_name = 'all'

    if not username or not password:
        return jsonify({'success': False, 'error': 'Username and password are required'}), 400
    if role == 'user' and not assigned_grades:
        return jsonify({
            'success': False,
            'error': 'Please choose at least one grade for this servant.'
        }), 400
    if role == 'sub_admin' and class_name == 'all':
        return jsonify({
            'success': False,
            'error': 'Please choose the class for this sub-admin.'
        }), 400

    try:
        result = db.add_user(username, password, role, class_name, assigned_grades, assigned_sections)
        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@role_required('admin')
def remove_user(user_id):
    """Delete a user."""
    try:
        success = db.delete_user(user_id)
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Cannot delete user (last admin or not found)'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>/password', methods=['PUT'])
@role_required('admin')
def change_password(user_id):
    """Change a user's password."""
    data = request.json
    new_password = data.get('password', '')

    if not new_password or len(new_password) < 4:
        return jsonify({'success': False, 'error': 'Password must be at least 4 characters'}), 400

    try:
        success = db.update_user_password(user_id, new_password)
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'User not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>/class', methods=['PUT'])
@role_required('admin')
def change_user_class(user_id):
    """Change a user's assigned class."""
    data = request.json or {}
    class_name = data.get('class_name', 'all')

    try:
        success = db.update_user_class(user_id, class_name)
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Invalid class or user not found'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>/grades', methods=['PUT'])
@role_required('admin')
def change_user_grades(user_id):
    """Change a user's assigned grades."""
    data = request.json or {}
    assigned_grades = data.get('assigned_grades', [])

    try:
        success = db.update_user_grades(user_id, assigned_grades)
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Invalid grades or user not found'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>/sections', methods=['PUT'])
@role_required('admin')
def change_user_sections(user_id):
    """Change a user's assigned boys/girls section."""
    data = request.json or {}
    assigned_sections = data.get('assigned_sections', [])

    try:
        success = db.update_user_sections(user_id, assigned_sections)
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Invalid section assignment or user not found'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== SERVANT REPORTS ====================

@app.route('/api/servant-report', methods=['POST'])
@login_required
def submit_servant_report():
    """Servant submits an attendance or eftikad report to notify admin."""
    data = request.json
    servant_username = data.get('servant_username', 'Unknown')
    report_type = data.get('report_type', 'attendance')  # 'attendance' or 'eftikad'
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    total = data.get('total', 0)
    present = data.get('present', 0)
    absent = data.get('absent', 0)
    notes = data.get('notes', '')

    try:
        report_id = db.submit_servant_report(
            servant_username, report_type, date, total, present, absent, notes
        )
        return jsonify({'success': True, 'report_id': report_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/servant-reports', methods=['GET'])
@min_role_required('sub_admin')
def get_servant_reports():
    """Get all servant reports, optionally filtered by date and/or type."""
    date_filter = request.args.get('date')
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    type_filter = request.args.get('type')
    grade_filter = request.args.get('grade')
    try:
        reports = db.get_servant_reports(date_filter, type_filter, date_from, date_to)
        grade_summary = db.get_grade_report_summary(date_from or date_filter, date_to or date_filter, grade_filter)
        return jsonify({'reports': reports, 'grade_summary': grade_summary})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/servant-reports/<int:report_id>', methods=['DELETE'])
@min_role_required('sub_admin')
def delete_servant_report(report_id):
    """Delete a servant report."""
    try:
        success = db.delete_servant_report(report_id)
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Report not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/servant-reports/mark-read', methods=['POST'])
@min_role_required('sub_admin')
def mark_reports_read():
    """Mark all reports as read (clears notification badge)."""
    try:
        db.mark_reports_read()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/servant-reports/unread-count', methods=['GET'])
@min_role_required('sub_admin')
def unread_count():
    """Get number of unread servant reports (for badge)."""
    try:
        count = db.get_unread_report_count()
        return jsonify({'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/servant-reports/download', methods=['GET'])
@min_role_required('sub_admin')
def download_servant_reports():
    """Download servant reports as an Excel file. Optional ?date=, ?from=, ?to=, and ?type= filters."""
    date_filter = request.args.get('date')
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    type_filter = request.args.get('type')
    grade_filter = request.args.get('grade')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import tempfile

        reports = db.get_servant_reports(date_filter, type_filter, date_from, date_to)
        grade_summary = db.get_grade_report_summary(date_from or date_filter, date_to or date_filter, grade_filter)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Servant Reports'

        # Header row
        headers = ['#', 'Servant', 'Type', 'Date', 'Total', 'Present', 'Absent', 'Notes', 'Submitted At']
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color='1F2535', end_color='1F2535', fill_type='solid')
        header_font = Font(bold=True, color='5B8EF0')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for i, r in enumerate(reports, start=1):
            rtype = '📋 Attendance' if r.get('report_type') == 'attendance' else '🏠 Eftikad Visit'
            ws.append([
                i,
                r.get('servant_username', ''),
                rtype,
                r.get('date', ''),
                r.get('total_count', 0),
                r.get('present_count', 0),
                r.get('absent_count', 0),
                r.get('notes', ''),
                r.get('submitted_at', '')
            ])

        summary_ws = wb.create_sheet('Grade Summary')
        summary_headers = [
            'Grade', 'Total Students', 'Present Students', 'Absent Students', 'Attendance Records',
            'Present Records', 'Absent Records', 'Rewarded Students', 'Reward Events',
            'Reward Points', 'Eftikad Students', 'Eftikad Visits'
        ]
        summary_ws.append(summary_headers)
        for cell in summary_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row in grade_summary:
            summary_ws.append([
                row.get('grade_label', ''),
                row.get('total_students', 0),
                row.get('present_students', 0),
                row.get('absent_students', 0),
                row.get('attendance_records', 0),
                row.get('present_count', 0),
                row.get('absent_count', 0),
                row.get('rewarded_students', 0),
                row.get('reward_events', 0),
                row.get('reward_points', 0),
                row.get('eftikad_students', 0),
                row.get('eftikad_visits', 0),
            ])

        for i, w in enumerate([16, 16, 18, 18, 18, 16, 16, 18, 14, 14, 16, 14], start=1):
            summary_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Column widths
        col_widths = [5, 20, 18, 14, 8, 10, 10, 30, 22]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Save temp
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        if date_filter:
            date_str = date_filter
        elif date_from or date_to:
            date_str = f"{date_from or 'start'}_to_{date_to or 'end'}"
        else:
            date_str = datetime.now().strftime('%Y%m%d')
        fname = f'servant_reports_{date_str}.xlsx'
        return send_file(tmp_path, as_attachment=True, download_name=fname)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("[INFO] Starting Church Attendance App...")
    print("[INFO] Access the app at: http://localhost:5000")
    print("[INFO] Admin dashboard at: http://localhost:5000/admin")
    app.run(debug=True, host='0.0.0.0', port=5000)
